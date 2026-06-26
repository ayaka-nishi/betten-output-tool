#!/usr/bin/env python3
"""
別添1（Word）・別添2（Excel）出力ツール（ローカルWebアプリ版）

このPC内だけで動くWebサーバーを起動し、自動でブラウザを開きます。
外部のネットワークには一切接続しません。

起動方法：
  python3 app.py
"""

import importlib.util
import io
import os
import socket
import sys
import tempfile
import threading
import uuid
import webbrowser
import zipfile
from collections import defaultdict

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook

# PyInstallerで.app/.exeに固めた場合、リソースは一時展開フォルダ(sys._MEIPASS)に入る
if getattr(sys, "frozen", False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SHEET_NAME  = "証明書まとめ"
SUBMIT_MARK = "○"

# 証明書作成.py を動的に読み込み、別添1/別添2 生成関数を再利用する
_spec = importlib.util.spec_from_file_location(
    "seimeisho_sakusei", os.path.join(BASE_DIR, "証明書作成.py")
)
seimeisho = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(seimeisho)

GOV_LABELS = [seimeisho.municipality_label(pref, city) for pref, city in seimeisho.MUNICIPALITIES]
GOV_MAP    = {seimeisho.municipality_label(pref, city): (pref, city)
              for pref, city in seimeisho.MUNICIPALITIES}

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
)

# 生成結果（ZIP）を一時保持する（同一PC内のみ・ブラウザ閉じたら不要になる想定）
_RESULTS: dict[str, bytes] = {}


def read_submitted_products(file_stream, area_default: str) -> dict:
    """J列（提出）が○の商品のみを {事業者名: [商品リスト]} で返す"""
    wb = load_workbook(file_stream, data_only=True)
    ws = wb[SHEET_NAME]

    companies: dict[str, list] = defaultdict(list)
    current_company = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells   = list(row) + [None] * 10
        company = cells[0]
        name    = cells[2]
        price   = cells[4]
        cost_b  = cells[5]
        area    = cells[6]
        retail  = cells[7]
        submit  = cells[9]

        if company and str(company).strip():
            current_company = str(company).strip()

        if not name or not current_company:
            continue
        if str(submit).strip() != SUBMIT_MARK:
            continue

        companies[current_company].append({
            "name":   str(name).strip(),
            "price":  price,
            "cost_b": cost_b,
            "area":   str(area).strip() if area else area_default,
            "retail": retail,
        })

    return {c: p for c, p in companies.items() if p}


@app.route("/")
def index():
    return render_template("index.html", gov_labels=GOV_LABELS)


@app.route("/gov_hint")
def gov_hint():
    """市長名の入力状況に応じて、別添にどう記載されるかをその場で返す"""
    gov_label  = request.args.get("gov_label", GOV_LABELS[0])
    mayor_name = request.args.get("mayor_name", "")
    if gov_label not in GOV_MAP:
        return jsonify({"error": "自治体が不正です"}), 400
    pref, city = GOV_MAP[gov_label]
    _, gov_head, _ = seimeisho.build_gov_names(pref, city, mayor_name)
    if mayor_name.strip():
        text = f"別添には「{gov_head}　殿」と記載されます"
    else:
        text = f"※市長名が空欄のため、別添には「{gov_head}　殿」と記載されます（氏名なし）"
    return jsonify({"text": text})


@app.route("/generate", methods=["POST"])
def generate():
    file = request.files.get("xlsx_file")
    if not file or not file.filename:
        return jsonify({"error": "① 対象ファイルを選択してください"}), 400

    gov_label  = request.form.get("gov_label", "")
    mayor_name = request.form.get("mayor_name", "")

    if gov_label not in GOV_MAP:
        return jsonify({"error": "自治体の選択が不正です"}), 400

    pref, city = GOV_MAP[gov_label]
    gov_name, gov_head, area_default = seimeisho.build_gov_names(pref, city, mayor_name)

    try:
        file_bytes = io.BytesIO(file.read())
        companies  = read_submitted_products(file_bytes, area_default=area_default)
    except Exception as e:
        return jsonify({"error": f"ファイルの読み込みに失敗しました: {e}"}), 400

    if not companies:
        return jsonify({
            "error": "提出（○）の商品が見つかりませんでした。J列に○がついているか確認してください。"
        }), 400

    total = sum(len(p) for p in companies.values())
    logs  = [f"自治体: {gov_name}　市長: {gov_head}",
             f"対象事業者: {len(companies)}社 / {total}件", ""]

    work_dir = os.path.join(tempfile.gettempdir(), f"betten_{uuid.uuid4().hex}")
    os.makedirs(work_dir, exist_ok=True)

    try:
        logs.append("【別添1 生成（Word）】")
        for company, products in companies.items():
            fname     = f"別添1_{seimeisho.safe_filename(company)}.docx"
            file_path = os.path.join(work_dir, fname)
            seimeisho.create_betten1_word(
                company, products, file_path, gov_name=gov_name, gov_head=gov_head)
            logs.append(f"✓ {fname}　（{len(products)}件）")

        logs.append("")
        logs.append("【別添2 生成（Excel）】")
        betten2_path = os.path.join(work_dir, "別添2.xlsx")
        seimeisho.create_betten2_excel(companies, betten2_path, gov_name=gov_name)
        logs.append(f"✓ 別添2.xlsx　（{len(companies)}社 / {total}件）")

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname in os.listdir(work_dir):
                zf.write(os.path.join(work_dir, fname), fname)
        zip_buffer.seek(0)

        token = uuid.uuid4().hex
        _RESULTS[token] = zip_buffer.getvalue()

        logs.append("")
        logs.append("完了しました！")

        return jsonify({
            "logs": logs,
            "download_token": token,
            "n_companies": len(companies),
            "n_products": total,
        })
    finally:
        for fname in os.listdir(work_dir):
            os.remove(os.path.join(work_dir, fname))
        os.rmdir(work_dir)


@app.route("/download/<token>")
def download(token):
    data = _RESULTS.pop(token, None)
    if data is None:
        return "ダウンロード期限が切れました。もう一度生成してください。", 404
    return send_file(
        io.BytesIO(data), mimetype="application/zip",
        as_attachment=True, download_name="別添1_別添2.zip",
    )


def _free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


if __name__ == "__main__":
    port = _free_port()
    url  = f"http://127.0.0.1:{port}"
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    print(f"\n別添出力ツールを起動しました: {url}")
    print("（このウィンドウを閉じるとアプリが終了します）\n")
    app.run(host="127.0.0.1", port=port, debug=False)
